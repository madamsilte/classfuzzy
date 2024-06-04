from openpyxl import load_workbook
from collections import namedtuple, defaultdict
import enum
import tqdm

# =============================================================================
# Arquitetura básica
# =============================================================================

Regra = namedtuple('Regra', 'proposicoes consequente')

class Niveis(enum.IntEnum):
    BAIXO = 0
    MEDIO = 1
    ALTO = 2

def fn_trapezio(x, a, b, c, d, alargamento=0.25):
    dx1 = (b - a) * alargamento
    dx2 = (d - c) * alargamento
    a -= dx1
    b += dx1
    c -= dx2
    d += dx2
    if x < a:
        return 0
    elif x < b:
        return (x - a) / (b - a)
    elif x < c:
        return 1
    elif x <= d:
        return (d-x) / (d-c)
    else:
        return 0    

def classificar(teores, regras, t_norm=min, fn_implicacao=min, fn_agregacao=max):
    # Nomenclatura conforme exemplo 2.5
    UFC = {}
    U = defaultdict(list)
    FR = defaultdict(list)
    # Passo 1: Fuzzificação, aplicação da T-Norma e construção dos conjuntos de saída
    for regra in regras:
        print(f'Litologia {regra.consequente[0]}:')
        mu_proposicoes = []
        for elemento, conjunto in regra.proposicoes:
            entrada = teores[elemento]
            fn = fn_entrada[elemento][conjunto]
            print(f'\t {elemento}={entrada} é {Niveis(conjunto).name}: {fn(entrada)}')
            mu_proposicoes.append(fn(entrada))
        mu_r = t_norm(mu_proposicoes)
        litologia, conjunto = regra.consequente
        fn = fn_saida[litologia][conjunto]
        fr_valores = []
        for i in range(1001):
            x = i * 0.1
            mu_saida = fn(x)
            #print(fn_implicacao(mu_r, mu_saida))
            fr_valores.append(fn_implicacao(mu_r, mu_saida))
        FR[litologia].append(fr_valores)
    # Passo 2: Agregação
    for litologia, conjuntos in FR.items():
        for i in range(1001):
            U[litologia].append(fn_agregacao([conjunto[i] for conjunto in conjuntos]))
    # Passo 3: Centro de Massa
    for litologia, conjunto in U.items():
        num = 0
        den = 0
        for i, y in enumerate(conjunto):
            x = i * 0.1
            num += x * y
            den += y
        UFC[litologia] = num/den if den != 0 else 0
    return UFC

# =============================================================================
# Lendo parâmetros
# =============================================================================

fn_entrada = defaultdict(list)
fn_saida = defaultdict(list)

wb = load_workbook('dados.xlsx')
ws = wb.worksheets[1]

# Funções de pertinência das variáveis de entrada
for i in range(3, ws.max_row+1):
    elemento = ws[i][0].value
    for offset in range(1, 10, 4):
        pontos = []
        for j in range(offset, offset+4):
            pontos.append(ws[i][j].value)
        if not any(pontos):
            break
        fn = lambda x,pontos=pontos: fn_trapezio(x, *pontos)
        fn_entrada[elemento].append(fn)

# Funções de pertinência das variáveis de saída
for litologia in ['CB', 'FO', 'NL', 'AL', 'FL', 'PI', 'CBMS', 'CBMG']:
    fn_saida[litologia].extend([
        lambda x: fn_trapezio(x, -10, -8, 20, 30),
        lambda x: fn_trapezio(x, 20, 30, 60, 70),
        lambda x: fn_trapezio(x, 70, 80, 100, 110)
    ])

# =============================================================================
# Teste
# =============================================================================\

regras = [
    Regra([('FE2O3', Niveis.BAIXO), ('AL2O3', Niveis.BAIXO), ('MGO', Niveis.ALTO), ('TIO2', Niveis.BAIXO)], 
        ('CB', Niveis.ALTO)
    ),
    Regra([('CAO', Niveis.MEDIO), ('P2O5', Niveis.ALTO), ('TIO2', Niveis.MEDIO), ('FE2O3', Niveis.ALTO), 
           ('AL2O3', Niveis.MEDIO)], 
        ('FO', Niveis.ALTO)
    ),
    Regra([('FE2O3', Niveis.ALTO), ('P2O5', Niveis.MEDIO), ('SIO2', Niveis.BAIXO), 
           ('AL2O3', Niveis.BAIXO), ('NB2O5', Niveis.ALTO)], 
        ('NL', Niveis.ALTO)
    ),
    Regra([('P2O5', Niveis.BAIXO), ('AL2O3', Niveis.ALTO), ('CAO', Niveis.BAIXO)],
        ('AL', Niveis.ALTO)
    ),
    Regra([('SIO2', Niveis.MEDIO), ('P2O5', Niveis.MEDIO), ('MGO', Niveis.MEDIO)], 
        ('FL', Niveis.ALTO)
    ),
    Regra([('AL2O3', Niveis.MEDIO), ('TIO2', Niveis.MEDIO), ('BAO', Niveis.BAIXO), ('MGO', Niveis.MEDIO), 
           ('FE2O3', Niveis.MEDIO)], 
        ('PI', Niveis.ALTO)
    ),
    Regra([('FE2O3', Niveis.MEDIO), ('MGO', Niveis.BAIXO), ('AL2O3', Niveis.MEDIO)], 
        ('CBMS', Niveis.ALTO)  
    ),
    Regra([('P2O5', Niveis.ALTO), ('NB2O5', Niveis.MEDIO)], 
        ('CBMG', Niveis.ALTO)     
    )
    ]

def teste1():
    exemplos = [
        {'FE2O3': 6.5, 'P2O5': 3, 'SIO2': 8, 'AL2O3': 0.5, }, # CB
        {'FE2O3': 31, 'P2O5': 7.9, 'SIO2': 14, 'AL2O3': 0.6, }, # FO
        {'FE2O3': 32, 'P2O5': 7, 'SIO2': 8, 'AL2O3': 0.5, }, # NL
        {'FE2O3': 16, 'P2O5': 4, 'SIO2': 11, 'AL2O3': 14, }, # AL
    ]

    for exemplo in exemplos:
        ufc = classificar(exemplo, regras)
        for litologia, confianca in ufc.items():
            print(litologia, confianca)
        print('='*80)

def teste2():
    with open('saida.csv', 'w', encoding='utf8') as fd:
        ws = wb.worksheets[0]
        cabecalho = [ws[1][j].value for j in range(1, ws.max_column)]
        # Funções de pertinência das variáveis de entrada
        print(','.join(fn_saida.keys()), file=fd)
        for i in tqdm.tqdm(range(2, ws.max_row+1)):
            teores = [ws[i][j].value for j in range(1, ws.max_column)]
            dic_teores = dict(zip(cabecalho, teores))
            ufc = classificar(dic_teores, regras)
            saidas = {}
            for litologia, confianca in ufc.items():
                saidas[litologia] = confianca
            print(','.join([f'{v}' for v in saidas.values()]), file=fd)

def teste3():
    elementos = ['NB2O5', 'P2O5', 'SIO2', 'FE2O3', 'BAO', 'CAO', 'MGO', 'TIO2', 'RCP', 'P2O5AP', 'AL2O3']
    teores = [0.1768, 0.897, 38.8756, 36.3982, 0.46, 0.4972, 1.7568, 4.6432, 0.5852, 0.3704, 0.14]
    exemplo = dict(zip(elementos, teores))
    ufc = classificar(exemplo, regras)
    for litologia, confianca in ufc.items():
        print(litologia, confianca)
        print('='*80)

teste3()
